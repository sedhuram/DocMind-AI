import csv
from dataclasses import dataclass
from pathlib import Path

import docx
import openpyxl
import pypdf


@dataclass
class ParsedPage:
    text: str
    page_number: int | None


class UnsupportedFileTypeError(Exception):
    pass


def parse_file(file_path: Path) -> list[ParsedPage]:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return _parse_pdf(file_path)
    if suffix == ".docx":
        return _parse_docx(file_path)
    if suffix in (".txt", ".md"):
        return _parse_text(file_path)
    if suffix == ".csv":
        return _parse_csv(file_path)
    if suffix == ".xlsx":
        return _parse_xlsx(file_path)
    raise UnsupportedFileTypeError(f"Unsupported file type: {suffix}")


def _parse_pdf(file_path: Path) -> list[ParsedPage]:
    reader = pypdf.PdfReader(str(file_path))
    pages: list[ParsedPage] = []
    for index, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(ParsedPage(text=text, page_number=index + 1))
    return pages


def _parse_docx(file_path: Path) -> list[ParsedPage]:
    document = docx.Document(str(file_path))
    text = "\n\n".join(p.text for p in document.paragraphs if p.text.strip())
    return [ParsedPage(text=text, page_number=None)] if text.strip() else []


def _parse_text(file_path: Path) -> list[ParsedPage]:
    text = file_path.read_text(encoding="utf-8", errors="ignore")
    return [ParsedPage(text=text, page_number=None)] if text.strip() else []


def _parse_csv(file_path: Path) -> list[ParsedPage]:
    lines = []
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            if any(cell.strip() for cell in row):
                lines.append(" | ".join(cell.strip() for cell in row))
    text = "\n".join(lines)
    return [ParsedPage(text=text, page_number=None)] if text.strip() else []


def _parse_xlsx(file_path: Path) -> list[ParsedPage]:
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    sheet_blocks = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                row_str = " | ".join(str(cell).strip() if cell is not None else "" for cell in row)
                rows.append(row_str)
        if rows:
            sheet_blocks.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    text = "\n\n".join(sheet_blocks)
    return [ParsedPage(text=text, page_number=None)] if text.strip() else []

